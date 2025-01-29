"""Utility to add a column with the color per row in a spreadsheet."""

import argparse

import openpyxl


def insert_color_column(input_path: str, output_path: str):
    """Adds a "color" column to a copy of the input spreadsheet file.

    Args:
        input_path: Path to the input Excel 2010 compatible file.
        output_path: Path at which to save the copy.
    """
    workbook = openpyxl.load_workbook(input_path)
    sheet = workbook.active
    # Note: indexes in the sheet are 1-based.
    output_column_index = sheet.max_column + 1
    for i, row in enumerate(sheet.rows):
        if i == 0:
            # Fill in the name for the new column
            sheet.cell(row=i + 1, column=output_column_index, value="color")
        else:
            # Take the first non-transparent color encountered, going
            # from the left
            color = ""
            for cell in row:
                cell_color = cell.fill.fgColor.rgb
                # rgb is sometimes not a string, perhaps when no fill has been set.
                # Skip these.
                if not isinstance(cell_color, str):
                    continue
                # Otherwise, if it's in the expected format (alpha + RGB), check
                # whether it's fully transparent, and if not, grab it.
                if len(cell_color) == 8 and not cell_color.startswith("00"):
                    color = cell_color[2:]
                    break
            if color:
                sheet.cell(row=i + 1, column=output_column_index, value=color)
    workbook.save(output_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        prog="add_xl_color_column",
        description="Add a column with the color found in each spreadsheet row.",
    )
    parser.add_argument("-i", "--input")
    parser.add_argument("-o", "--output")
    args = parser.parse_args()
    insert_color_column(input_path=args.input, output_path=args.output)
